import json
import os
import shutil
import hashlib
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


INPUT_XLSX = os.path.join("input", "istoki.xlsx")
DIST_DIR = "dist"
DOCS_DIR = "docs"

SONGS_JSON_NAME = "songs.json"
LATEST_JSON_NAME = "latest.json"

PAGES_BASE = "https://aleyakim.github.io/istoki-catalog/"
SONGS_URL = PAGES_BASE + SONGS_JSON_NAME  # https://aleyakim.github.io/istoki-catalog/songs.json


def _s(v: Any) -> str:
    """Cell value -> trimmed string, None -> ''."""
    if v is None:
        return ""
    return str(v).strip()


def _split_list(cell: Any) -> List[str]:
    """Split by ';' into trimmed non-empty items."""
    raw = _s(cell)
    if not raw:
        return []
    parts = [p.strip() for p in raw.split(";")]
    return [p for p in parts if p]


def _opt(cell: Any) -> Optional[str]:
    """Optional string: '' -> None"""
    t = _s(cell)
    return t if t else None


def _now_iso_utc() -> str:
    # e.g. 2026-02-09T12:34:56Z
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _fail(msg: str) -> None:
    raise SystemExit(f"ERROR: {msg}")


def _warn(msg: str) -> None:
    print(f"WARNING: {msg}")


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _read_file_bytes(path: str) -> Optional[bytes]:
    if not os.path.exists(path):
        return None
    with open(path, "rb") as f:
        return f.read()


def read_meta(ws) -> Dict[str, str]:
    meta: Dict[str, str] = {}
    # Expect columns: A=key, B=value; from row 1..N
    for row in ws.iter_rows(min_row=1, values_only=True):
        key = _s(row[0])
        val = _s(row[1]) if len(row) > 1 else ""
        if key:
            meta[key] = val
    return meta


def sheet_headers(ws) -> Dict[str, int]:
    # First row is headers
    headers: Dict[str, int] = {}
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, cell in enumerate(first):
        name = _s(cell)
        if name:
            headers[name] = idx
    return headers


def get_cell(row: Tuple[Any, ...], headers: Dict[str, int], name: str) -> Any:
    idx = headers.get(name)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def build_song_dict(row: Tuple[Any, ...], h: Dict[str, int]) -> Dict[str, Any]:
    # Required
    song_id = _s(get_cell(row, h, "id"))
    title = _s(get_cell(row, h, "title"))
    hub = _s(get_cell(row, h, "hub"))
    lyrics = _s(get_cell(row, h, "lyrics"))

    if not song_id:
        _fail("songs: empty id")
    if not title:
        _fail(f"songs: empty title for id={song_id}")
    if not hub:
        _fail(f"songs: empty hub for id={song_id}")
    if not lyrics:
        _fail(f"songs: empty lyrics for id={song_id}")

    song = {
        "id": song_id,
        "title": title,
        "altTitles": _s(get_cell(row, h, "altTitles")),
        "hub": hub,
        "regionDetail": _s(get_cell(row, h, "regionDetail")),
        "genre": _s(get_cell(row, h, "genre")),
        "themes": _split_list(get_cell(row, h, "themes")),
        "cossackHost": _s(get_cell(row, h, "cossackHost")),
        "languageTag": _s(get_cell(row, h, "languageTag")) or "русский",
        "contextShort": _s(get_cell(row, h, "contextShort")),
        "lyrics": lyrics,
        "glossary": [],   # filled later
        "sourceNote": _s(get_cell(row, h, "sourceNote")),
        "audioUrl": _s(get_cell(row, h, "audioUrl")),
        "minusUrl": _s(get_cell(row, h, "minusUrl")),
        "videoUrl": _s(get_cell(row, h, "videoUrl")),
        "rightsStatus": _s(get_cell(row, h, "rightsStatus")) or "NONE",
        "externalLinks": _split_list(get_cell(row, h, "externalLinks")),
        "versions": [],   # filled later
    }
    return song


def build_version_dict(row: Tuple[Any, ...], h: Dict[str, int]) -> Tuple[str, Dict[str, Any]]:
    song_id = _s(get_cell(row, h, "songId"))
    ver_id = _s(get_cell(row, h, "id"))

    if not song_id:
        _fail("versions: empty songId")
    if not ver_id:
        _fail(f"versions: empty id for songId={song_id}")

    # SongVersion fields are nullable in Kotlin -> omit keys when empty
    ver: Dict[str, Any] = {"id": ver_id}
    for k in ["title", "lyrics", "sourceNote", "audioUrl", "minusUrl", "videoUrl"]:
        v = _opt(get_cell(row, h, k))
        if v is not None:
            ver[k] = v

    return song_id, ver


def build_glossary_item(row: Tuple[Any, ...], h: Dict[str, int]) -> Tuple[str, Dict[str, Any]]:
    song_id = _s(get_cell(row, h, "songId"))
    term = _s(get_cell(row, h, "term"))
    definition = _s(get_cell(row, h, "definition"))

    if not song_id:
        _fail("glossary: empty songId")
    if not term:
        _fail(f"glossary: empty term for songId={song_id}")
    if not definition:
        _fail(f"glossary: empty definition for songId={song_id}, term={term}")

    return song_id, {"term": term, "definition": definition}


def main() -> None:
    if not os.path.exists(INPUT_XLSX):
        _fail(f"Input file not found: {INPUT_XLSX}")

    # data_only=True: tries to read calculated values instead of formulas
    wb = load_workbook(INPUT_XLSX, data_only=True)

    for sheet in ["meta", "songs", "versions", "glossary"]:
        if sheet not in wb.sheetnames:
            _fail(f"Missing sheet '{sheet}' in {INPUT_XLSX}")

    meta = read_meta(wb["meta"])

    if "catalogVersion" not in meta:
        _fail("meta: missing catalogVersion")
    try:
        catalog_version = int(meta["catalogVersion"])
    except ValueError:
        _fail(f"meta: catalogVersion is not int: {meta['catalogVersion']}")

    base_media_url = meta.get("baseMediaUrl", "").strip()

    # ---- songs ----
    ws_songs = wb["songs"]
    hs = sheet_headers(ws_songs)
    required_cols = ["id", "title", "hub", "lyrics"]
    for col in required_cols:
        if col not in hs:
            _fail(f"songs: missing required column '{col}'")

    songs_by_id: Dict[str, Dict[str, Any]] = {}
    for row_idx, row in enumerate(ws_songs.iter_rows(min_row=2, values_only=True), start=2):
        # skip totally empty rows
        if all((_s(c) == "" for c in row)):
            continue

        song = build_song_dict(row, hs)
        song_id = song["id"]
        if song_id in songs_by_id:
            _fail(f"songs: duplicate id '{song_id}' (row {row_idx})")
        songs_by_id[song_id] = song

    if not songs_by_id:
        _fail("songs: no songs found")

    # ---- versions ----
    ws_versions = wb["versions"]
    hv = sheet_headers(ws_versions)
    if "songId" not in hv or "id" not in hv:
        _fail("versions: required columns are 'songId' and 'id'")

    version_ids_per_song: Dict[str, set] = {}
    for row_idx, row in enumerate(ws_versions.iter_rows(min_row=2, values_only=True), start=2):
        if all((_s(c) == "" for c in row)):
            continue
        song_id, ver = build_version_dict(row, hv)

        if song_id not in songs_by_id:
            _fail(f"versions: unknown songId '{song_id}' (row {row_idx})")

        seen = version_ids_per_song.setdefault(song_id, set())
        if ver["id"] in seen:
            _fail(f"versions: duplicate version id '{ver['id']}' for songId={song_id} (row {row_idx})")
        seen.add(ver["id"])

        songs_by_id[song_id]["versions"].append(ver)

    # ---- glossary ----
    ws_glossary = wb["glossary"]
    hg = sheet_headers(ws_glossary)
    for col in ["songId", "term", "definition"]:
        if col not in hg:
            _fail(f"glossary: missing required column '{col}'")

    term_defs_global: Dict[str, Dict[str, List[str]]] = {}  # termNorm -> def -> [songIds]
    terms_per_song: Dict[str, set] = {}

    for row_idx, row in enumerate(ws_glossary.iter_rows(min_row=2, values_only=True), start=2):
        if all((_s(c) == "" for c in row)):
            continue
        song_id, item = build_glossary_item(row, hg)

        if song_id not in songs_by_id:
            _fail(f"glossary: unknown songId '{song_id}' (row {row_idx})")

        term = item["term"]
        definition = item["definition"]

        # per-song duplicate term = error
        seen_terms = terms_per_song.setdefault(song_id, set())
        term_norm = term.strip().lower()
        if term_norm in seen_terms:
            _fail(f"glossary: duplicate term '{term}' within songId={song_id} (row {row_idx})")
        seen_terms.add(term_norm)

        songs_by_id[song_id]["glossary"].append(item)

        # global warning: same term has different definitions across songs
        def_norm = definition.strip()
        term_defs_global.setdefault(term_norm, {}).setdefault(def_norm, []).append(song_id)

    for term_norm, defs_map in term_defs_global.items():
        if len(defs_map) > 1:
            _warn(f"term '{term_norm}' has {len(defs_map)} different definitions across songs:")
            for defn, song_ids in defs_map.items():
                preview = defn.replace("\n", " ")[:120]
                _warn(f"  used in {sorted(set(song_ids))}: {preview}")

    # stable order
    songs: List[Dict[str, Any]] = [songs_by_id[k] for k in sorted(songs_by_id.keys())]

    # Prepare new songs.json bytes (deterministic) for writing + comparisons
    new_songs_json_bytes = json.dumps(songs, ensure_ascii=False, indent=2).encode("utf-8")

    # --- strict versioning gate (enabled in CI) ---
    strict = os.getenv("STRICT_VERSIONING", "").strip().lower() in ("1", "true", "yes", "on")

    old_latest_bytes = _read_file_bytes(os.path.join(DOCS_DIR, LATEST_JSON_NAME))
    old_songs_bytes = _read_file_bytes(os.path.join(DOCS_DIR, SONGS_JSON_NAME))

    if strict and old_latest_bytes is not None:
        try:
            old_manifest = json.loads(old_latest_bytes.decode("utf-8"))
            old_version = int(old_manifest.get("catalogVersion", 0))
            old_base_media_url = str(old_manifest.get("baseMediaUrl", "")).strip()
        except Exception as e:
            _fail(f"STRICT_VERSIONING: cannot parse existing docs/latest.json: {e}")

        songs_changed = False
        if old_songs_bytes is not None:
            try:
                # Parse old JSON and re-dump with same formatting to avoid false diff from line endings
                old_obj = json.loads(old_songs_bytes.decode("utf-8"))
                old_norm_bytes = json.dumps(old_obj, ensure_ascii=False, indent=2).encode("utf-8")
                songs_changed = _sha256_bytes(old_norm_bytes) != _sha256_bytes(new_songs_json_bytes)
            except Exception as e:
                _fail(f"STRICT_VERSIONING: cannot parse existing docs/songs.json: {e}")

        base_media_changed = old_base_media_url != base_media_url.strip()

        if (songs_changed or base_media_changed) and catalog_version <= old_version:
            details = []
            if songs_changed:
                details.append("songs.json changed")
            if base_media_changed:
                details.append(f"baseMediaUrl changed ('{old_base_media_url}' -> '{base_media_url.strip()}')")

            _fail(
                "STRICT_VERSIONING: content changed but catalogVersion was not bumped. "
                f"old={old_version}, new={catalog_version}. Причина: {', '.join(details)}. "
                "Открой input/istoki.xlsx → sheet meta → catalogVersion и увеличь (например +1), затем commit/push."
            )

    os.makedirs(DIST_DIR, exist_ok=True)
    os.makedirs(DOCS_DIR, exist_ok=True)

    songs_path_dist = os.path.join(DIST_DIR, SONGS_JSON_NAME)
    latest_path_dist = os.path.join(DIST_DIR, LATEST_JSON_NAME)

    # Write songs.json (bytes)
    with open(songs_path_dist, "wb") as f:
        f.write(new_songs_json_bytes)

    manifest = {
        "catalogVersion": catalog_version,
        "publishedAt": _now_iso_utc(),
        "songsUrl": SONGS_URL,
        "baseMediaUrl": base_media_url,
    }

    with open(latest_path_dist, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    # Copy to docs/ (what Pages serves)
    shutil.copyfile(songs_path_dist, os.path.join(DOCS_DIR, SONGS_JSON_NAME))
    shutil.copyfile(latest_path_dist, os.path.join(DOCS_DIR, LATEST_JSON_NAME))

    print("OK: generated:")
    print(f"  {os.path.join(DOCS_DIR, SONGS_JSON_NAME)}  (songs: {len(songs)})")
    print(f"  {os.path.join(DOCS_DIR, LATEST_JSON_NAME)}  (catalogVersion: {catalog_version})")
    if base_media_url:
        print(f"  baseMediaUrl: {base_media_url}")
    else:
        _warn("baseMediaUrl is empty in meta sheet")


if __name__ == "__main__":
    main()